import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn.linear_model import LinearRegression
import io
import base64
import warnings
from openpyxl import Workbook  # ✅ AGREGAR ESTA IMPORTACIÓN
warnings.filterwarnings('ignore')

# Configuración optimizada para producción
st.set_page_config(
    page_title="PRISMA - Proyección y Simulación para Metas COMGES",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Constantes
MONTHS = ['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO',
          'JULIO','AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']
M2N = {m:i+1 for i,m in enumerate(MONTHS)}
N2M = {i+1:m for i,m in enumerate(MONTHS)}
N_SIM = 10000
SEED = 42

def read_and_process_file(uploaded_file):
    """Lee y procesa archivos CSV o Excel subidos"""
    try:
        # Determinar el tipo de archivo
        file_extension = uploaded_file.name.split('.')[-1].lower()
        
        if file_extension in ['csv']:
            return _process_csv_file(uploaded_file)
        elif file_extension in ['xls', 'xlsx']:
            return _process_excel_file(uploaded_file)
        else:
            st.error(f"Formato de archivo no soportado: {file_extension}")
            return None
            
    except Exception as e:
        st.error(f"Error al leer el archivo: {e}")
        return None

def _process_csv_file(uploaded_file):
    """Procesa archivos CSV"""
    encodings = ['latin-1', 'utf-8', 'cp1252', 'iso-8859-1']
    separators = [';', ',', '\t']
    
    for encoding in encodings:
        for separator in separators:
            try:
                uploaded_file.seek(0)
                df = pd.read_csv(
                    uploaded_file, 
                    sep=separator,
                    encoding=encoding,
                    decimal=',',
                    engine='python',
                    thousands='.'
                )
                st.success(f"✅ CSV leído con: encoding={encoding}, separator='{separator}'")
                return _process_dataframe(df)
            except Exception as e:
                continue
    
    st.error("No se pudo leer el archivo CSV. Verifica el formato.")
    return None

def _process_excel_file(uploaded_file):
    """Procesa archivos Excel"""
    try:
        uploaded_file.seek(0)
        excel_file = pd.ExcelFile(uploaded_file)
        
        if len(excel_file.sheet_names) > 1:
            st.info(f"📑 Hojas disponibles: {excel_file.sheet_names}")
            sheet_name = excel_file.sheet_names[0]
            st.write(f"Usando hoja: '{sheet_name}'")
        else:
            sheet_name = excel_file.sheet_names[0]
        
        # Determinar engine según extensión
        engine = 'openpyxl' if uploaded_file.name.endswith('.xlsx') else 'xlrd'
        
        df = pd.read_excel(uploaded_file, sheet_name=sheet_name, engine=engine)
        st.success(f"✅ Excel leído correctamente - Hoja: '{sheet_name}'")
        return _process_dataframe(df)
        
    except Exception as e:
        st.error(f"Error al leer archivo Excel: {e}")
        return None

def _process_dataframe(df):
    """Procesa el DataFrame independientemente del origen"""
    st.write("📋 **Columnas detectadas en el archivo:**")
    st.write(df.columns.tolist())
    
    # Normalizar nombres de columnas
    column_mapping = {}
    for col in df.columns:
        col_clean = str(col).strip().upper()
        
        if any(x in col_clean for x in ['AČO', 'AÑO', 'ANIO', 'YEAR']):
            column_mapping[col] = 'ANIO'
        elif any(x in col_clean for x in ['MES', 'MONTH']):
            column_mapping[col] = 'MES'
        elif any(x in col_clean for x in ['NUMERADOR', 'NUMERATOR', 'CASOS', 'EVENTOS']):
            column_mapping[col] = 'NUMERADOR'
        elif any(x in col_clean for x in ['DENOMINADOR', 'DENOMINATOR', 'POBLACION', 'TOTAL']):
            column_mapping[col] = 'DENOMINADOR'
        elif '%' in col_clean or 'PORCENTAJE' in col_clean:
            continue
    
    df = df.rename(columns=column_mapping)
    
    # Verificar columnas requeridas
    required_columns = ['ANIO', 'MES', 'NUMERADOR', 'DENOMINADOR']
    missing_columns = [col for col in required_columns if col not in df.columns]
    
    if missing_columns:
        st.error(f"❌ Faltan columnas requeridas: {missing_columns}")
        st.write("Columnas disponibles:", df.columns.tolist())
        return None
    
    # Limpieza y procesamiento
    df["MES"] = df["MES"].astype(str).str.strip().str.upper()
    
    mes_mapping = {
        'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4, 'MAYO': 5, 'JUNIO': 6,
        'JULIO': 7, 'AGOSTO': 8, 'SEPTIEMBRE': 9, 'OCTUBRE': 10, 'NOVIEMBRE': 11, 'DICIEMBRE': 12,
        'ENE': 1, 'FEB': 2, 'MAR': 3, 'ABR': 4, 'MAY': 5, 'JUN': 6,
        'JUL': 7, 'AGO': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DIC': 12
    }
    
    df["m"] = df["MES"].map(mes_mapping)
    
    # Convertir columnas numéricas
    for c in ["NUMERADOR", "DENOMINADOR"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    
    # Eliminar filas problemáticas
    df = df.dropna(how="all").sort_values(["ANIO", "m"]).reset_index(drop=True)
    essential = ["ANIO", "m", "NUMERADOR", "DENOMINADOR"]
    df = df.dropna(subset=essential).reset_index(drop=True)
    
    if len(df) == 0:
        st.error("❌ No hay datos válidos después del procesamiento")
        return None
    
    df["pct"] = df["NUMERADOR"] / df["DENOMINADOR"]
    st.success(f"✅ Datos procesados correctamente: {len(df)} registros válidos")
    
    return df

# ============================================================================
# MÉTODOS DE SIMULACIÓN MEJORADOS - AHORA DEVUELVEN NUM, DEN, PCT
# ============================================================================

def average_method(df):
    """Método del promedio - retorna numerador, denominador y porcentaje PROMEDIO mensual"""
    # Usar los últimos 3 meses para el cálculo del promedio
    last = df.tail(3)
    
    # Calcular promedios mensuales correctamente
    avg_num = last["NUMERADOR"].mean()
    avg_den = last["DENOMINADOR"].mean()
    
    # Si no hay datos suficientes, usar todos los datos
    if not np.isfinite(avg_num) or not np.isfinite(avg_den) or avg_den <= 0:
        avg_num = df["NUMERADOR"].mean()
        avg_den = df["DENOMINADOR"].mean()
    
    if avg_den <= 0:
        return 0.0, 0.0, 0.0
    
    # Calcular porcentaje basado en los promedios
    avg_pct = avg_num / avg_den
    return avg_num, avg_den, avg_pct


def linear_trend_forecast_improved(y, steps):
    """Pronóstico lineal mejorado con manejo de errores"""
    y_clean = pd.to_numeric(y, errors="coerce").dropna()
    
    if len(y_clean) < 2:
        last_val = y_clean.iloc[-1] if len(y_clean) > 0 else 0.0
        return np.full(steps, last_val)
    
    x = np.arange(len(y_clean)).reshape(-1, 1)
    model = LinearRegression().fit(x, y_clean.values)
    xf = np.arange(len(y_clean), len(y_clean) + steps).reshape(-1, 1)
    predictions = model.predict(xf)
    
    return np.maximum(predictions, 0)

def seasonal_indices(df):
    """Calcula índices estacionales"""
    if len(df) < 6:
        return pd.Series(dtype=float), pd.Series(dtype=float)
    
    muN, muD = df["NUMERADOR"].mean(), df["DENOMINADOR"].mean()
    
    if muN <= 0 or muD <= 0:
        return pd.Series(dtype=float), pd.Series(dtype=float)
    
    idxN = df.groupby("m")["NUMERADOR"].mean() / muN
    idxD = df.groupby("m")["DENOMINADOR"].mean() / muD
    
    return idxN, idxD

def mc_simulation_adaptive(muN, sdN, muD, sdD, n=N_SIM, seed=SEED):
    """Simulación Monte Carlo adaptativa - retorna num, den, pct"""
    rng = np.random.default_rng(seed)
    
    # Determinar distribución para NUMERADOR
    if muN < 20:
        N = rng.poisson(muN, size=n)
    elif muN >= 30 and sdN**2 <= muN * 1.5:
        N = rng.normal(muN, sdN, size=n)
        N = np.maximum(N, 0)
    else:
        if sdN <= 0:
            sdN = np.sqrt(muN)
        k = (muN / sdN) ** 2
        theta = sdN ** 2 / muN
        N = rng.gamma(max(k, 0.1), max(theta, 1e-6), size=n)
    
    # Determinar distribución para DENOMINADOR
    if muD < 20:
        D = rng.poisson(muD, size=n)
    elif muD >= 30 and sdD**2 <= muD * 1.5:
        D = rng.normal(muD, sdD, size=n)
        D = np.maximum(D, 1)
    else:
        if sdD <= 0:
            sdD = np.sqrt(muD)
        k = (muD / sdD) ** 2
        theta = sdD ** 2 / muD
        D = rng.gamma(max(k, 0.1), max(theta, 1e-6), size=n)
    
    D = np.clip(D, 1e-6, None)
    
    # Calcular medianas
    median_N = float(np.median(N))
    median_D = float(np.median(D))
    median_pct = median_N / median_D if median_D > 0 else 0.0
    
    return median_N, median_D, median_pct

def next_three_months_from_last(df):
    """Determina los próximos períodos a proyectar según la periodicidad"""
    try:
        periodo_info = detect_periodicity(df)
        
        if periodo_info['tipo'] == 'mensual':
            last_m = int(df["m"].iloc[-1])
            months = [((last_m + i - 1) % 12) + 1 for i in (1,2,3)]
            return months
        elif periodo_info['tipo'] in ['trimestral', 'semestral']:
            # Para trimestral/semestral, proyectar solo 1 período pero representado como lista
            last_m = int(df["m"].iloc[-1])
            # Calcular el próximo período
            if periodo_info['tipo'] == 'trimestral':
                next_month = ((last_m + 2) // 3) * 3 + 1  # Siguiente trimestre
                if next_month > 12:
                    next_month = 1
            else:  # semestral
                next_month = 7 if last_m <= 6 else 1  # Siguiente semestre
            
            return [next_month]  # Retornar como lista con un elemento
        else:
            # Por defecto, mensual (3 meses)
            last_m = int(df["m"].iloc[-1])
            months = [((last_m + i - 1) % 12) + 1 for i in (1,2,3)]
            return months
    except Exception as e:
        # Fallback a mensual si hay error
        last_m = int(df["m"].iloc[-1])
        months = [((last_m + i - 1) % 12) + 1 for i in (1,2,3)]
        return months

def generate_comprehensive_report(df, results):
    """Genera reporte con todos los resultados"""
    
    # Crear tabla resumen de porcentajes
    pct_table = pd.DataFrame({
        'Mes': [r['Mes'] for r in results],
        'Promedio': [r['Promedio_pct'] for r in results],
        'Pronostico_Lineal': [r['Pronostico_Lineal_pct'] for r in results],
        'MC_Adaptativo': [r['MC_Adaptativo_pct'] for r in results],
        'MC_Adaptativo_Estacional': [r['MC_Adaptativo_Estacional_pct'] for r in results]
    })
    
    html_content = f"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Reporte Completo de Proyecciones</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }}
            .container {{ max-width: 100%; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; }}
            h1 {{ color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 10px; }}
            h2 {{ color: #34495e; margin-top: 30px; }}
            .table-wrapper {{ overflow-x: auto; margin: 20px 0; }}
            table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
            th, td {{ padding: 8px; text-align: center; border: 1px solid #ddd; }}
            th {{ background: #3498db; color: white; }}
            .section {{ margin: 30px 0; }}
            .method-card {{ background: #f8f9fa; padding: 15px; margin: 10px 0; border-left: 4px solid #3498db; }}
            .highlight {{ background: #e8f4f8; }}
        </style>
    </head>
    <body>
        <div class="container">
            <h1>📊 Reporte Completo de Proyecciones</h1>
            
            <div class="section">
                <h2>📈 Resumen de Proyecciones - Porcentajes</h2>
                {pct_table.to_html(index=False, float_format=lambda x: f'{x*100:.2f}%')}
            </div>
            
            <div class="section">
                <h2>🔍 Detalle por Método y Mes</h2>
    """
    
    # Agregar detalles por mes
    for i, result in enumerate(results):
        html_content += f"""
                <div class="method-card">
                    <h3>📅 {result['Mes']}</h3>
                    <table>
                        <tr>
                            <th>Método</th>
                            <th>Numerador</th>
                            <th>Denominador</th>
                            <th>Porcentaje</th>
                        </tr>
                        <tr>
                            <td><strong>Promedio</strong></td>
                            <td>{result['Promedio_num']:.1f}</td>
                            <td>{result['Promedio_den']:.1f}</td>
                            <td><strong>{result['Promedio_pct']*100:.2f}%</strong></td>
                        </tr>
                        <tr>
                            <td><strong>Pronóstico Lineal</strong></td>
                            <td>{result['Pronostico_Lineal_num']:.1f}</td>
                            <td>{result['Pronostico_Lineal_den']:.1f}</td>
                            <td><strong>{result['Pronostico_Lineal_pct']*100:.2f}%</strong></td>
                        </tr>
                        <tr class="highlight">
                            <td><strong>MC Adaptativo</strong></td>
                            <td>{result['MC_Adaptativo_num']:.1f}</td>
                            <td>{result['MC_Adaptativo_den']:.1f}</td>
                            <td><strong>{result['MC_Adaptativo_pct']*100:.2f}%</strong></td>
                        </tr>
                        <tr class="highlight">
                            <td><strong>MC Adaptativo Estacional</strong></td>
                            <td>{result['MC_Adaptativo_Estacional_num']:.1f}</td>
                            <td>{result['MC_Adaptativo_Estacional_den']:.1f}</td>
                            <td><strong>{result['MC_Adaptativo_Estacional_pct']*100:.2f}%</strong></td>
                        </tr>
                    </table>
                </div>
        """
    
    html_content += """
            </div>
            
            <div class="section">
                <h2>💡 Métodos Utilizados</h2>
                <div class="method-card">
                    <h4>📊 Promedio</h4>
                    <p>Basado en los últimos 3 meses históricos. Método conservador.</p>
                </div>
                <div class="method-card">
                    <h4>📈 Pronóstico Lineal</h4>
                    <p>Extrapolación de tendencia usando regresión lineal simple.</p>
                </div>
                <div class="method-card">
                    <h4>🎲 Monte Carlo Adaptativo</h4>
                    <p>Simulación probabilística que selecciona automáticamente la distribución más apropiada según las características de los datos.</p>
                </div>
                <div class="method-card">
                    <h4>🌐 Monte Carlo Adaptativo Estacional</h4>
                    <p>Incluye ajuste por patrones estacionales mensuales históricos.</p>
                </div>
            </div>
            
            <div class="section" style="background: #d5f4e6; padding: 20px; border-radius: 5px;">
                <h2>✅ Recomendación</h2>
                <p><strong>Método recomendado: Monte Carlo Adaptativo Estacional</strong></p>
                <p>Este método combina la flexibilidad de la simulación Monte Carlo con el ajuste estacional, proporcionando las proyecciones más robustas para datos epidemiológicos.</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    return html_content

def detect_periodicity(df):
    """
    Detecta si los datos son mensuales, trimestrales o semestrales
    basándose en la frecuencia de los registros
    """
    if len(df) < 2:
        return {'tipo': 'mensual', 'confianza': 'baja', 'mensaje': 'Datos insuficientes para detectar periodicidad'}
    
    try:
        # Ordenar por año y mes
        df_sorted = df.sort_values(['ANIO', 'm']).reset_index(drop=True)
        
        # Calcular diferencias entre meses consecutivos
        diffs = []
        for i in range(1, len(df_sorted)):
            current = df_sorted.iloc[i]
            previous = df_sorted.iloc[i-1]
            
            # Calcular diferencia en meses
            diff_meses = (current['ANIO'] - previous['ANIO']) * 12 + (current['m'] - previous['m'])
            diffs.append(diff_meses)
        
        # Si no hay diferencias calculadas, retornar mensual por defecto
        if not diffs:
            return {'tipo': 'mensual', 'confianza': 'baja', 'mensaje': 'No se pudieron calcular diferencias'}
        
        # Analizar diferencias
        diff_unique = list(set(diffs))
        diff_counts = {diff: diffs.count(diff) for diff in diff_unique}
        
        # Determinar periodicidad
        if all(diff == 1 for diff in diffs):
            return {'tipo': 'mensual', 'confianza': 'alta', 'diferencia': 1}
        elif all(diff == 3 for diff in diffs):
            return {'tipo': 'trimestral', 'confianza': 'alta', 'diferencia': 3}
        elif all(diff == 6 for diff in diffs):
            return {'tipo': 'semestral', 'confianza': 'alta', 'diferencia': 6}
        else:
            # Periodicidad mixta o irregular
            most_common_diff = max(diff_counts, key=diff_counts.get)
            if diff_counts[most_common_diff] / len(diffs) >= 0.7:  # 70% de coincidencia
                tipo = {1: 'mensual', 3: 'trimestral', 6: 'semestral'}.get(most_common_diff, 'irregular')
                return {'tipo': tipo, 'confianza': 'media', 'diferencia': most_common_diff}
            else:
                return {'tipo': 'irregular', 'confianza': 'baja', 'diferencia': diff_unique}
    
    except Exception as e:
        # En caso de cualquier error, retornar mensual por defecto
        return {'tipo': 'mensual', 'confianza': 'baja', 'mensaje': f'Error en detección: {str(e)}'}
    
def get_next_period(df, period_type):
    """
    Calcula el próximo período según el tipo de periodicidad
    """
    last_row = df.sort_values(['ANIO', 'm']).iloc[-1]
    last_year = last_row['ANIO']
    last_month = last_row['m']
    
    if period_type == 'mensual':
        next_month = (last_month % 12) + 1
        next_year = last_year if next_month > last_month else last_year + 1
        return {
            'nombre': f"{N2M[next_month]} {next_year}",
            'mes_num': next_month,
            'año': next_year
        }
    elif period_type == 'trimestral':
        # Encontrar el trimestre actual (1: Ene-Mar, 2: Abr-Jun, etc.)
        current_quarter = (last_month - 1) // 3 + 1
        next_quarter = (current_quarter % 4) + 1
        next_year = last_year if next_quarter > current_quarter else last_year + 1
        
        quarter_months = {
            1: ('ENERO', 'MARZO'),
            2: ('ABRIL', 'JUNIO'), 
            3: ('JULIO', 'SEPTIEMBRE'),
            4: ('OCTUBRE', 'DICIEMBRE')
        }
        mes_inicio, mes_fin = quarter_months[next_quarter]
        return {
            'nombre': f"Trimestre {next_quarter} ({mes_inicio}-{mes_fin}) {next_year}",
            'trimestre': next_quarter,
            'año': next_year
        }
    elif period_type == 'semestral':
        # Encontrar el semestre actual (1: Ene-Jun, 2: Jul-Dic)
        current_semester = 1 if last_month <= 6 else 2
        next_semester = (current_semester % 2) + 1
        next_year = last_year if next_semester > current_semester else last_year + 1
        
        semester_range = {
            1: ('ENERO', 'JUNIO'),
            2: ('JULIO', 'DICIEMBRE')
        }
        mes_inicio, mes_fin = semester_range[next_semester]
        return {
            'nombre': f"Semestre {next_semester} ({mes_inicio}-{mes_fin}) {next_year}",
            'semestre': next_semester,
            'año': next_year
        }
    else:
        # Por defecto, mensual
        next_month = (last_month % 12) + 1
        next_year = last_year if next_month > last_month else last_year + 1
        return {
            'nombre': f"{N2M[next_month]} {next_year}",
            'mes_num': next_month,
            'año': next_year
        }

#Simulacion de porcentaje
# Agregar esta función después de las otras funciones de simulación
def simulate_target_percentage(df, target_pct):
    """
    Simula numeradores y denominadores necesarios para alcanzar un porcentaje objetivo
    """
    # Detectar periodicidad
    periodo_info = detect_periodicity(df)
    
    # Obtener el próximo período según la periodicidad
    future_periods = next_three_months_from_last(df)
    next_period_num = future_periods[0]
    
    # Determinar nombre del período
    if periodo_info['tipo'] == 'trimestral':
        period_name = f"Trimestre {(next_period_num - 1) // 3 + 1}"
    elif periodo_info['tipo'] == 'semestral':
        period_name = "Semestre 1" if next_period_num <= 6 else "Semestre 2"
    else:
        period_name = N2M[next_period_num]
    
    # Parámetros globales
    muN, sdN = df["NUMERADOR"].mean(), df["NUMERADOR"].std(ddof=1)
    muD, sdD = df["DENOMINADOR"].mean(), df["DENOMINADOR"].std(ddof=1)
    
    # Índices estacionales (solo para mensual)
    idxN, idxD = seasonal_indices(df) if periodo_info['tipo'] == 'mensual' else (pd.Series(), pd.Series())
    
    # Método 1: Promedio - usar promedio mensual histórico
    avg_num, avg_den, avg_pct = average_method(df)
    target_avg_num = max(0, round(target_pct * avg_den))
    target_avg_den = max(1, round(avg_den))
    
    # Método 2: Lineal - proyectar y redondear
    steps = 3 if periodo_info['tipo'] == 'mensual' else 1
    predN_lin = linear_trend_forecast_improved(df["NUMERADOR"], steps=steps)[0]
    predD_lin = linear_trend_forecast_improved(df["DENOMINADOR"], steps=steps)[0]
    target_lin_num = max(0, round(target_pct * predD_lin))
    target_lin_den = max(1, round(predD_lin))
    
    # Método 3: MC Adaptativo - redondear resultados
    mc_num, mc_den, mc_pct = mc_simulation_adaptive(muN, sdN, muD, sdD, seed=SEED)
    target_mc_num = max(0, round(target_pct * mc_den))
    target_mc_den = max(1, round(mc_den))
    
    # Método 4: MC Estacional (solo para mensual)
    if periodo_info['tipo'] == 'mensual' and not idxN.empty:
        muNm = muN * idxN.get(next_period_num, 1.0)
        muDm = muD * idxD.get(next_period_num, 1.0)
        mc_seas_num, mc_seas_den, mc_seas_pct = mc_simulation_adaptive(muNm, sdN, muDm, sdD, seed=SEED + 1000)
        target_mc_seas_num = max(0, round(target_pct * mc_seas_den))
        target_mc_seas_den = max(1, round(mc_seas_den))
    else:
        # Para trimestral/semestral, usar MC normal
        target_mc_seas_num = target_mc_num
        target_mc_seas_den = target_mc_den
    
    return {
        'Mes': period_name,
        'Tipo_Periodo': periodo_info['tipo'],
        'Porcentaje_Objetivo': target_pct,
        # Promedio
        'Promedio_num': target_avg_num,
        'Promedio_den': target_avg_den,
        'Promedio_pct': target_pct,
        # Lineal
        'Pronostico_Lineal_num': target_lin_num,
        'Pronostico_Lineal_den': target_lin_den,
        'Pronostico_Lineal_pct': target_pct,
        # MC Adaptativo
        'MC_Adaptativo_num': target_mc_num,
        'MC_Adaptativo_den': target_mc_den,
        'MC_Adaptativo_pct': target_pct,
        # MC Adaptativo Estacional
        'MC_Adaptativo_Estacional_num': target_mc_seas_num,
        'MC_Adaptativo_Estacional_den': target_mc_seas_den,
        'MC_Adaptativo_Estacional_pct': target_pct
    }

def generate_simulation_report(df, simulation_results):
    """Genera reporte HTML para la simulación de metas"""
    
    html_content = f"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Simulación de Meta - PRISMA</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }}
            .container {{ max-width: 100%; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; }}
            h1 {{ color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 10px; }}
            h2 {{ color: #34495e; margin-top: 30px; }}
            .table-wrapper {{ overflow-x: auto; margin: 20px 0; }}
            table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
            th, td {{ padding: 8px; text-align: center; border: 1px solid #ddd; }}
            th {{ background: #3498db; color: white; }}
            .section {{ margin: 30px 0; }}
            .method-card {{ background: #f8f9fa; padding: 15px; margin: 10px 0; border-left: 4px solid #3498db; }}
            .highlight {{ background: #e8f4f8; }}
            .target-info {{ background: #d5f4e6; padding: 20px; border-radius: 5px; margin: 20px 0; }}
        </style>
    </head>
    <body>
        <div class="container">
            <h1>🎯 Simulación de Meta - PRISMA</h1>
            
            <div class="target-info">
                <h2>📋 Información de la Meta</h2>
                <p><strong>Mes objetivo:</strong> {simulation_results['Mes']}</p>
                <p><strong>Porcentaje deseado:</strong> {simulation_results['Porcentaje_Objetivo']*100:.2f}%</p>
                <p><strong>Interpretación:</strong> Esta simulación muestra los valores de numerador y denominador necesarios para alcanzar la meta del {simulation_results['Porcentaje_Objetivo']*100:.2f}% según cada método de proyección.</p>
            </div>

            <div class="section">
                <h2>📊 Valores Requeridos por Método</h2>
                <table>
                    <tr>
                        <th>Método</th>
                        <th>Numerador Requerido</th>
                        <th>Denominador Proyectado</th>
                        <th>Porcentaje Resultante</th>
                        <th>Diferencia Numérica</th>
                    </tr>
                    <tr>
                        <td><strong>Promedio</strong></td>
                        <td>{simulation_results['Promedio_num']:.1f}</td>
                        <td>{simulation_results['Promedio_den']:.1f}</td>
                        <td><strong>{simulation_results['Promedio_pct']*100:.2f}%</strong></td>
                        <td>{simulation_results['Promedio_num'] - df['NUMERADOR'].mean():.1f}</td>
                    </tr>
                    <tr>
                        <td><strong>Pronóstico Lineal</strong></td>
                        <td>{simulation_results['Pronostico_Lineal_num']:.1f}</td>
                        <td>{simulation_results['Pronostico_Lineal_den']:.1f}</td>
                        <td><strong>{simulation_results['Pronostico_Lineal_pct']*100:.2f}%</strong></td>
                        <td>{simulation_results['Pronostico_Lineal_num'] - df['NUMERADOR'].mean():.1f}</td>
                    </tr>
                    <tr class="highlight">
                        <td><strong>MC Adaptativo</strong></td>
                        <td>{simulation_results['MC_Adaptativo_num']:.1f}</td>
                        <td>{simulation_results['MC_Adaptativo_den']:.1f}</td>
                        <td><strong>{simulation_results['MC_Adaptativo_pct']*100:.2f}%</strong></td>
                        <td>{simulation_results['MC_Adaptativo_num'] - df['NUMERADOR'].mean():.1f}</td>
                    </tr>
                    <tr class="highlight">
                        <td><strong>MC Adaptativo Estacional</strong></td>
                        <td>{simulation_results['MC_Adaptativo_Estacional_num']:.1f}</td>
                        <td>{simulation_results['MC_Adaptativo_Estacional_den']:.1f}</td>
                        <td><strong>{simulation_results['MC_Adaptativo_Estacional_pct']*100:.2f}%</strong></td>
                        <td>{simulation_results['MC_Adaptativo_Estacional_num'] - df['NUMERADOR'].mean():.1f}</td>
                    </tr>
                </table>
            </div>

            <div class="section">
                <h2>💡 Recomendaciones por Método</h2>
                <div class="method-card">
                    <h4>📊 Promedio</h4>
                    <p>Para alcanzar el {simulation_results['Porcentaje_Objetivo']*100:.2f}%, necesitarías <strong>{simulation_results['Promedio_num']:.1f} eventos</strong> manteniendo el denominador histórico de {simulation_results['Promedio_den']:.1f}.</p>
                </div>
                <div class="method-card">
                    <h4>📈 Pronóstico Lineal</h4>
                    <p>Considerando la tendencia, necesitarías <strong>{simulation_results['Pronostico_Lineal_num']:.1f} eventos</strong> con un denominador proyectado de {simulation_results['Pronostico_Lineal_den']:.1f}.</p>
                </div>
                <div class="method-card highlight">
                    <h4>🎲 MC Adaptativo (Recomendado)</h4>
                    <p>Basado en la variabilidad histórica, se necesitan <strong>{simulation_results['MC_Adaptativo_num']:.1f} eventos</strong> con un denominador estimado de {simulation_results['MC_Adaptativo_den']:.1f}.</p>
                </div>
                <div class="method-card highlight">
                    <h4>🌐 MC Adaptativo Estacional</h4>
                    <p>Considerando patrones mensuales, se requieren <strong>{simulation_results['MC_Adaptativo_Estacional_num']:.1f} eventos</strong> con denominador de {simulation_results['MC_Adaptativo_Estacional_den']:.1f}.</p>
                </div>
            </div>

            <div class="section" style="background: #e8f4f8; padding: 20px; border-radius: 5px;">
                <h2>📈 Resumen Ejecutivo</h2>
                <p><strong>Meta:</strong> {simulation_results['Porcentaje_Objetivo']*100:.2f}% en {simulation_results['Mes']}</p>
                <p><strong>Rango de numeradores requeridos:</strong> {min([simulation_results['Promedio_num'], simulation_results['Pronostico_Lineal_num'], simulation_results['MC_Adaptativo_num'], simulation_results['MC_Adaptativo_Estacional_num']]):.1f} - {max([simulation_results['Promedio_num'], simulation_results['Pronostico_Lineal_num'], simulation_results['MC_Adaptativo_num'], simulation_results['MC_Adaptativo_Estacional_num']]):.1f} eventos</p>
                <p><strong>Recomendación principal:</strong> Planificar para <strong>{simulation_results['MC_Adaptativo_Estacional_num']:.1f} eventos</strong> (método MC Estacional)</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    return html_content


def main():
    st.title("📊 PRISMA - Proyección y Simulación para Metas COMGES")
    st.markdown("""
    **Sistema de simulación de proyecciones estadisticas**  
    Carga archivos CSV o Excel para generar proyecciones detalladas de los próximos 3 meses.
    """)
    
    # Sidebar
    # En la función main(), dentro del with st.sidebar:
    with st.sidebar:
        st.header("📁 Cargar Datos")
        uploaded_file = st.file_uploader(
            "Selecciona tu archivo de datos", 
            type=['csv', 'xlsx', 'xls'],
            help="Formatos soportados: CSV, Excel (.xlsx, .xls). Columnas requeridas: AÑO, MES, NUMERADOR, DENOMINADOR"
        )
        
        st.markdown("---")
        st.header("📋 Plantillas")
        
        # Plantilla para CSV
        st.subheader("📄 Plantilla CSV")
        
        # Crear datos de ejemplo para la plantilla (DEFINIR template_df ANTES de usarlo)
        template_data = {
            'AÑO': [2024, 2024, 2024, 2024, 2024, 2024],
            'MES': ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO'],
            'NUMERADOR': [150, 120, 180, 160, 140, 170],
            'DENOMINADOR': [1000, 950, 1100, 1050, 980, 1200]
        }
        
        template_df = pd.DataFrame(template_data)
        
        # Convertir a CSV
        csv_template = template_df.to_csv(index=False, sep=';', encoding='utf-8-sig')
        
        st.download_button(
            label="📥 Descargar Plantilla CSV",
            data=csv_template,
            file_name="plantilla_prisma.csv",
            mime="text/csv",
            help="Plantilla en formato CSV con separador punto y coma"
        )
        
        # Plantilla para Excel
        st.subheader("📊 Plantilla Excel")
        
        # Crear archivo Excel en memoria con manejo de errores
        try:
            excel_buffer = io.BytesIO()
            
            # Crear workbook desde cero en lugar de usar pd.ExcelWriter
            wb = Workbook()
            
            # Remover la hoja por defecto
            wb.remove(wb.active)
            
            # Crear hoja de datos
            ws_data = wb.create_sheet("Datos_Ejemplo")
            
            # Escribir encabezados
            headers = ['AÑO', 'MES', 'NUMERADOR', 'DENOMINADOR']
            for col, header in enumerate(headers, 1):
                ws_data.cell(row=1, column=col, value=header)
            
            # Escribir datos de ejemplo
            for row, (_, data_row) in enumerate(template_df.iterrows(), 2):
                ws_data.cell(row=row, column=1, value=data_row['AÑO'])
                ws_data.cell(row=row, column=2, value=data_row['MES'])
                ws_data.cell(row=row, column=3, value=data_row['NUMERADOR'])
                ws_data.cell(row=row, column=4, value=data_row['DENOMINADOR'])
            
            # Crear hoja de instrucciones
            ws_instructions = wb.create_sheet("Instrucciones")
            
            instructions_data = [
                ['Campo', 'Descripción', 'Ejemplo', 'Requerido'],
                ['AÑO', 'Año del registro', '2024, 2025', 'Sí'],
                ['MES', 'Nombre del mes en español en MAYÚSCULAS', 'ENERO, FEBRERO, MARZO', 'Sí'],
                ['NUMERADOR', 'Cantidad de casos/eventos (número)', '150, 120.5', 'Sí'],
                ['DENOMINADOR', 'Población/total (número)', '1000, 950.5', 'Sí'],
                ['', '', '', ''],
                ['Recomendaciones:', '', '', ''],
                ['- Mínimo 6 meses de datos', '', '', ''],
                ['- No incluir columna de porcentaje (%)', '', '', ''],
                ['- Formato consistente en todos los registros', '', '', '']
            ]
            
            for row, instruction_row in enumerate(instructions_data, 1):
                for col, value in enumerate(instruction_row, 1):
                    ws_instructions.cell(row=row, column=col, value=value)
            
            # Establecer la primera hoja como activa
            wb.active = ws_data
            
            # Guardar workbook en el buffer
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            st.download_button(
                label="📥 Descargar Plantilla Excel",
                data=excel_buffer,
                file_name="plantilla_prisma.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Plantilla en formato Excel con hojas de datos e instrucciones"
            )
            
        except Exception as e:
            st.error(f"Error al crear plantilla Excel: {e}")
            # Ofrecer alternativa simple
            simple_excel_buffer = io.BytesIO()
            template_df.to_excel(simple_excel_buffer, index=False, engine='openpyxl')
            simple_excel_buffer.seek(0)
            
            st.download_button(
                label="📥 Descargar Plantilla Simple Excel",
                data=simple_excel_buffer,
                file_name="plantilla_prisma_simple.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Plantilla básica en formato Excel"
            )
        
        # Información sobre el formato
        with st.expander("ℹ️ Instrucciones de formato"):
            st.markdown("""
            **📝 Formato requerido para los archivos:**
            
            **Columnas obligatorias:**
            - `AÑO`: Año del registro (ej: 2024)
            - `MES`: Nombre del mes en español (ej: ENERO, FEBRERO)
            - `NUMERADOR`: Valor numérico (ej: casos, eventos)
            - `DENOMINADOR`: Valor numérico (ej: población, total)
            
            **📋 Ejemplo de datos:**
            ```
            AÑO;MES;NUMERADOR;DENOMINADOR
            2024;ENERO;150;1000
            2024;FEBRERO;120;950
            2024;MARZO;180;1100
            ```
            
            **✅ Recomendaciones:**
            - Mínimo 6 meses de datos históricos
            - Meses en MAYÚSCULAS y nombres completos
            - Sin caracteres especiales en los números
            - No incluir la columna de porcentaje (%)
            """)
        
        st.markdown("---")
        # ... el resto del sidebar permanece igual
        st.header("🔄 Métodos de Proyección")
        st.markdown("""
        - **📊 Promedio**: Baseline histórico
        - **📈 Lineal**: Tendencia por regresión  
        - **🎲 MC Adaptativo**: Simulación probabilística
        - **🌐 MC Estacional**: Con ajuste mensual
        """)
        
        st.markdown("---")
        st.header("📈 Resultados Incluyen")
        st.markdown("""
        - Porcentajes proyectados
        - Numeradores estimados
        - Denominadores estimados
        - Comparación entre métodos
        """)

        st.markdown("---")
        st.markdown("""
        PRISMA - Proyección y Simulación para Metas COMGES  
        Versión 0.87 - 2025-06-15 
        © Christian Fuentes + IA.
                    
                    """)

    if uploaded_file is not None:
        # Procesar datos
        with st.spinner("Procesando datos..."):
            df = read_and_process_file(uploaded_file)
            
            if df is None or len(df) == 0:
                st.error("No se pudieron procesar los datos.")
                return

            # Mostrar información básica
            st.success(f"✅ {len(df)} registros procesados correctamente")

        # Realizar proyecciones
        with st.spinner("Calculando proyecciones..."):
            future_months = next_three_months_from_last(df)
            
            # Mostrar información de periodicidad
            periodo_info = detect_periodicity(df)
            if periodo_info['tipo'] == 'mensual':
                st.info(f"📅 **Periodicidad:** MENSUAL - Proyectando 3 meses")
            elif periodo_info['tipo'] == 'trimestral':
                st.info(f"📅 **Periodicidad:** TRIMESTRAL - Proyectando 1 trimestre")
            elif periodo_info['tipo'] == 'semestral':
                st.info(f"📅 **Periodicidad:** SEMESTRAL - Proyectando 1 semestre")
            else:
                st.warning(f"📅 **Periodicidad:** IRREGULAR - Proyectando 3 meses por defecto")
            
            # Parámetros globales
            muN, sdN = df["NUMERADOR"].mean(), df["NUMERADOR"].std(ddof=1)
            muD, sdD = df["DENOMINADOR"].mean(), df["DENOMINADOR"].std(ddof=1)
            
            # Índices estacionales (solo para mensual)
            idxN, idxD = seasonal_indices(df) if periodo_info['tipo'] == 'mensual' else (pd.Series(), pd.Series())
            
            # Método del promedio (constante para todos los períodos)
            avg_num, avg_den, avg_pct = average_method(df)
            
            # Métodos lineales - ajustar steps según periodicidad
            if periodo_info['tipo'] == 'mensual':
                steps = 3
            else:
                steps = 1
            
            predN_lin = linear_trend_forecast_improved(df["NUMERADOR"], steps=steps)
            predD_lin = linear_trend_forecast_improved(df["DENOMINADOR"], steps=steps)
            
            # Recolectar resultados
            results = []
            
            for i, month_num in enumerate(future_months):
                # Para periodicidades no mensuales, ajustar el nombre del período
                if periodo_info['tipo'] == 'trimestral':
                    period_name = f"Trimestre {(month_num - 1) // 3 + 1}"
                elif periodo_info['tipo'] == 'semestral':
                    period_name = "Semestre 1" if month_num <= 6 else "Semestre 2"
                else:
                    period_name = N2M[month_num]
                
                # Parámetros estacionales (solo para mensual)
                if not idxN.empty and periodo_info['tipo'] == 'mensual':
                    muNm = muN * idxN.get(month_num, 1.0)
                    muDm = muD * idxD.get(month_num, 1.0)
                    sdNm = sdN * idxN.get(month_num, 1.0)
                    sdDm = sdD * idxD.get(month_num, 1.0)
                else:
                    muNm, muDm, sdNm, sdDm = muN, muD, sdN, sdD
                
                # Método lineal
                lin_num = predN_lin[i]
                lin_den = predD_lin[i]
                lin_pct = lin_num / lin_den if lin_den > 0 else 0.0
                
                # Métodos Monte Carlo
                mc_num, mc_den, mc_pct = mc_simulation_adaptive(muN, sdN, muD, sdD, seed=SEED + i*100)
                
                # MC Estacional solo para datos mensuales
                if periodo_info['tipo'] == 'mensual' and not idxN.empty:
                    mc_seas_num, mc_seas_den, mc_seas_pct = mc_simulation_adaptive(muNm, sdNm, muDm, sdDm, seed=SEED + 1000 + i*100)
                else:
                    mc_seas_num, mc_seas_den, mc_seas_pct = mc_num, mc_den, mc_pct
                
                results.append({
                    'Mes': period_name,
                    # Promedio
                    'Promedio_num': avg_num,
                    'Promedio_den': avg_den,
                    'Promedio_pct': avg_pct,
                    # Lineal
                    'Pronostico_Lineal_num': lin_num,
                    'Pronostico_Lineal_den': lin_den,
                    'Pronostico_Lineal_pct': lin_pct,
                    # MC Adaptativo
                    'MC_Adaptativo_num': mc_num,
                    'MC_Adaptativo_den': mc_den,
                    'MC_Adaptativo_pct': mc_pct,
                    # MC Adaptativo Estacional
                    'MC_Adaptativo_Estacional_num': mc_seas_num,
                    'MC_Adaptativo_Estacional_den': mc_seas_den,
                    'MC_Adaptativo_Estacional_pct': mc_seas_pct
                })

        # Mostrar resultados en pestañas
        tab1, tab2, tab3, tab4, tab5 = st.tabs(["📊 Resultados Completos", "📈 Gráficos", "🔢 Datos Detallados", "🌐 Reporte HTML", "🎯 Simulador de Metas"])

        with tab1:
            # Mostrar título según periodicidad
            periodo_info = detect_periodicity(df)
            if periodo_info['tipo'] == 'mensual':
                st.subheader("Proyecciones Completas - Próximos 3 Meses")
            elif periodo_info['tipo'] == 'trimestral':
                st.subheader("Proyección - Próximo Trimestre")
            elif periodo_info['tipo'] == 'semestral':
                st.subheader("Proyección - Próximo Semestre")
            else:
                st.subheader("Proyecciones Completas - Todos los Métodos")
            
            # Crear DataFrame para display
            display_data = []
            # En la pestaña de resultados, modificar la creación de display_data:
            for result in results:
                for method in ['Promedio', 'Pronostico_Lineal', 'MC_Adaptativo', 'MC_Adaptativo_Estacional']:
                    display_data.append({
                        'Período': result['Mes'],  # Cambiar 'Mes' por 'Período'
                        'Método': method.replace('_', ' ').title(),
                        'Numerador': round(result[f'{method}_num']),  # Redondear valores
                        'Denominador': round(result[f'{method}_den']),  # Redondear valores
                        'Porcentaje': f"{result[f'{method}_pct']*100:.2f}%"
                    })
            
            display_df = pd.DataFrame(display_data)
            st.dataframe(display_df, use_container_width=True, hide_index=True)
            
            # Resumen de porcentajes
            st.subheader("Resumen de Porcentajes por Método")
            pct_summary = pd.DataFrame({
                'Mes': [r['Mes'] for r in results],
                'Promedio': [f"{r['Promedio_pct']*100:.2f}%" for r in results],
                'Pronóstico Lineal': [f"{r['Pronostico_Lineal_pct']*100:.2f}%" for r in results],
                'MC Adaptativo': [f"{r['MC_Adaptativo_pct']*100:.2f}%" for r in results],
                'MC Adaptativo Estacional': [f"{r['MC_Adaptativo_Estacional_pct']*100:.2f}%" for r in results]
            })
            st.dataframe(pct_summary, use_container_width=True, hide_index=True)
        
        with tab2:
            st.subheader("Comparación Gráfica de Métodos")
            
            # Gráfico de porcentajes. Fix
            fig, ax = plt.subplots(figsize=(12, 6))
            methods = ['Promedio', 'Pronostico_Lineal', 'MC_Adaptativo', 'MC_Adaptativo_Estacional']
            labels = ['Promedio', 'Lineal', 'MC Adaptativo', 'MC Estacional']
            colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728']
            
            x = np.arange(len(results))
            width = 0.2
            
            for i, (method, label, color) in enumerate(zip(methods, labels, colors)):
                percentages = [r[f'{method}_pct'] * 100 for r in results]
                bars = ax.bar(x + (i-1.5)*width, percentages, width, label=label, color=color, alpha=0.8)
                
                # Agregar valores en las barras
                for j, pct in enumerate(percentages):
                    ax.text(x[j] + (i-1.5)*width, pct + 0.1, f'{pct:.1f}%', 
                        ha='center', va='bottom', fontsize=9, fontweight='bold')
            
            ax.set_xticks(x)
            ax.set_xticklabels([r['Mes'] for r in results])
            ax.set_ylabel('Porcentaje (%)')
            ax.set_title('Comparación de Porcentajes Proyectados por Método')
            ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
            ax.grid(axis='y', alpha=0.3)
            plt.tight_layout()
            
            # ✅ MOSTRAR EL GRÁFICO
            st.pyplot(fig)
            
            # Análisis de variabilidad entre métodos
            st.subheader("📈 Análisis de Variabilidad entre Métodos")
            variability_data = []
            for r in results:
                pcts = [r['Promedio_pct']*100, r['Pronostico_Lineal_pct']*100, 
                    r['MC_Adaptativo_pct']*100, r['MC_Adaptativo_Estacional_pct']*100]
                variability_data.append({
                    'Mes': r['Mes'],
                    'Mínimo': f"{min(pcts):.2f}%",
                    'Máximo': f"{max(pcts):.2f}%", 
                    'Diferencia': f"{max(pcts)-min(pcts):.2f}%",
                    'Recomendado (MC Estacional)': f"{r['MC_Adaptativo_Estacional_pct']*100:.2f}%"
                })
            
            variability_df = pd.DataFrame(variability_data)
            st.dataframe(variability_df, use_container_width=True, hide_index=True)
            
            # Interpretación
            st.info("""
            **💡 Interpretación del Gráfico:**
            - **Promedio**: Método conservador basado en datos históricos recientes
            - **Lineal**: Captura tendencias ascendentes o descendentes
            - **MC Adaptativo**: Considera la variabilidad natural de los datos
            - **MC Estacional**: Incluye patrones mensuales históricos (Recomendado)
            """)
        
                    
        with tab3:
            st.subheader("Datos de Entrada Procesados")
            st.dataframe(df, use_container_width=True, hide_index=True)
            
            # Estadísticas descriptivas
            st.subheader("Estadísticas Descriptivas")
            col1, col2 = st.columns(2)
            
            with col1:
                st.write("**Numerador**")
                stats_n = df['NUMERADOR'].describe().reset_index()
                stats_n.columns = ['Estadística', 'Valor']
                # ✅ CORREGIDO: Ocultar índice
                st.dataframe(stats_n, use_container_width=True, hide_index=True)
    
            with col2:
                st.write("**Denominador**")
                stats_d = df['DENOMINADOR'].describe().reset_index()
                stats_d.columns = ['Estadística', 'Valor']
                # ✅ CORREGIDO: Ocultar índice
                st.dataframe(stats_d, use_container_width=True, hide_index=True)
        
        with tab4:
            st.subheader("Reporte HTML Completo")
            
            # Generar y mostrar reporte HTML
            html_content = generate_comprehensive_report(df, results)
            st.components.v1.html(html_content, height=1000, scrolling=True)
            
            # Botones de descarga
            col1, col2 = st.columns(2)
            
            with col1:
                # Descargar HTML
                st.download_button(
                    label="📥 Descargar Reporte HTML",
                    data=html_content,
                    file_name="reporte_proyecciones_completo.html",
                    mime="text/html"
                )
            
            with col2:
                # Descargar CSV con resultados
                results_csv = pd.DataFrame(results)
                csv_data = results_csv.to_csv(index=False, encoding='utf-8-sig')
                st.download_button(
                    label="📊 Descargar Resultados CSV",
                    data=csv_data,
                    file_name="resultados_proyecciones.csv",
                    mime="text/csv"
                )
        
        with tab5:
            st.subheader("🎯 Simulador de Metas")
            st.markdown("""
            **Establece un porcentaje objetivo y descubre los valores necesarios para alcanzarlo**  
            Esta herramienta calcula los numeradores y denominadores requeridos según diferentes métodos de proyección.
            """)
            
            # Input del usuario
            col1, col2 = st.columns([2, 1])
            
            with col1:
                target_percentage = st.slider(
                    "Porcentaje objetivo deseado:",
                    min_value=0.0,
                    max_value=100.0,
                    value=5.0,
                    step=0.1,
                    format="%.1f%%",
                    help="Selecciona el porcentaje de cumplimiento que deseas alcanzar"
                ) / 100.0  # Convertir a decimal
            
            with col2:
                st.metric("Meta establecida", f"{target_percentage*100:.1f}%")
            
            # Información contextual
            current_avg = df['pct'].mean() * 100
            st.info(f"📊 **Contexto histórico:** El porcentaje promedio actual es del {current_avg:.2f}%")
            
            if st.button("🚀 Calcular Valores Requeridos", type="primary"):
                with st.spinner("Calculando valores necesarios..."):
                    # Ejecutar simulación
                    simulation_results = simulate_target_percentage(df, target_percentage)
                    
                    # Mostrar resultados principales
                    st.success("✅ Simulación completada")
                    
                    # Métricas principales
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        st.metric("Mes Objetivo", simulation_results['Mes'])
                    with col2:
                        st.metric("Meta", f"{target_percentage*100:.1f}%")
                    with col3:
                        min_num = min([simulation_results['Promedio_num'], 
                                    simulation_results['Pronostico_Lineal_num'],
                                    simulation_results['MC_Adaptativo_num'],
                                    simulation_results['MC_Adaptativo_Estacional_num']])
                        st.metric("Mín. Numerador", f"{min_num:.1f}")
                    with col4:
                        max_num = max([simulation_results['Promedio_num'], 
                                    simulation_results['Pronostico_Lineal_num'],
                                    simulation_results['MC_Adaptativo_num'],
                                    simulation_results['MC_Adaptativo_Estacional_num']])
                        st.metric("Máx. Numerador", f"{max_num:.1f}")
                    
                    # Tabla de resultados
                    st.subheader("📋 Valores Requeridos por Método")
                    
                    simulation_data = []
                    methods = ['Promedio', 'Pronostico_Lineal', 'MC_Adaptativo', 'MC_Adaptativo_Estacional']
                    labels = ['Promedio', 'Lineal', 'MC Adaptativo', 'MC Estacional']
                    
                    for method, label in zip(methods, labels):
                        simulation_data.append({
                            'Método': label,
                            'Numerador Requerido': f"{simulation_results[f'{method}_num']:.1f}",
                            'Denominador Proyectado': f"{simulation_results[f'{method}_den']:.1f}",
                            'Porcentaje': f"{simulation_results[f'{method}_pct']*100:.2f}%",
                            'Incremento vs Promedio': f"{(simulation_results[f'{method}_num'] - df['NUMERADOR'].mean()):.1f}"
                        })
                    
                    simulation_df = pd.DataFrame(simulation_data)
                    st.dataframe(simulation_df, use_container_width=True, hide_index=True)
                    
                    # Gráfico comparativo
                    st.subheader("📈 Comparación de Numeradores Requeridos")
                    
                    fig, ax = plt.subplots(figsize=(10, 6))
                    methods = ['Promedio', 'Pronostico_Lineal', 'MC_Adaptativo', 'MC_Adaptativo_Estacional']
                    labels = ['Promedio', 'Lineal', 'MC Adaptativo', 'MC Estacional']
                    colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728']
                    
                    x = np.arange(len(methods))
                    numerators = [simulation_results[f'{method}_num'] for method in methods]
                    
                    bars = ax.bar(x, numerators, color=colors, alpha=0.8)
                    
                    # Línea del promedio histórico
                    historical_avg = df['NUMERADOR'].mean()
                    ax.axhline(y=historical_avg, color='red', linestyle='--', linewidth=2, 
                            label=f'Promedio histórico: {historical_avg:.1f}')
                    
                    # Etiquetas en las barras
                    for i, (bar, num) in enumerate(zip(bars, numerators)):
                        ax.text(bar.get_x() + bar.get_width()/2, num + max(numerators)*0.01, 
                            f'{num:.1f}', ha='center', va='bottom', fontsize=11, fontweight='bold')
                    
                    ax.set_xticks(x)
                    ax.set_xticklabels(labels)
                    ax.set_ylabel('Numerador Requerido')
                    ax.set_title(f'Numeradores Necesarios para Alcanzar {target_percentage*100:.1f}%')
                    ax.legend()
                    ax.grid(axis='y', alpha=0.3)
                    
                    st.pyplot(fig)
                    
                    # Análisis de brecha
                    st.subheader("🔍 Análisis de Brecha")
                    current_avg_num = df['NUMERADOR'].mean()
                    recommended_num = simulation_results['MC_Adaptativo_Estacional_num']
                    gap = recommended_num - current_avg_num
                    
                    if gap > 0:
                        st.warning(f"📈 **Necesitas incrementar** en aproximadamente {gap:.1f} eventos respecto al promedio histórico")
                    else:
                        st.success(f"📉 **Estás por encima de la meta** por aproximadamente {abs(gap):.1f} eventos")
                    
                    # Reporte HTML para descargar
                    st.subheader("🌐 Reporte de Simulación")
                    simulation_html = generate_simulation_report(df, simulation_results)
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        st.download_button(
                            label="📥 Descargar Reporte de Simulación",
                            data=simulation_html,
                            file_name=f"simulacion_meta_{target_percentage*100:.1f}percent.html",
                            mime="text/html"
                        )
                    
                    with col2:
                        # Datos para CSV
                        simulation_csv = pd.DataFrame([{
                            'Mes': simulation_results['Mes'],
                            'Porcentaje_Objetivo': simulation_results['Porcentaje_Objetivo'],
                            'Metodo': 'Promedio',
                            'Numerador_Requerido': simulation_results['Promedio_num'],
                            'Denominador_Proyectado': simulation_results['Promedio_den']
                        }, {
                            'Mes': simulation_results['Mes'],
                            'Porcentaje_Objetivo': simulation_results['Porcentaje_Objetivo'],
                            'Metodo': 'Lineal',
                            'Numerador_Requerido': simulation_results['Pronostico_Lineal_num'],
                            'Denominador_Proyectado': simulation_results['Pronostico_Lineal_den']
                        }, {
                            'Mes': simulation_results['Mes'],
                            'Porcentaje_Objetivo': simulation_results['Porcentaje_Objetivo'],
                            'Metodo': 'MC_Adaptativo',
                            'Numerador_Requerido': simulation_results['MC_Adaptativo_num'],
                            'Denominador_Proyectado': simulation_results['MC_Adaptativo_den']
                        }, {
                            'Mes': simulation_results['Mes'],
                            'Porcentaje_Objetivo': simulation_results['Porcentaje_Objetivo'],
                            'Metodo': 'MC_Estacional',
                            'Numerador_Requerido': simulation_results['MC_Adaptativo_Estacional_num'],
                            'Denominador_Proyectado': simulation_results['MC_Adaptativo_Estacional_den']
                        }])
                        
                        csv_data = simulation_csv.to_csv(index=False, encoding='utf-8-sig')
                        st.download_button(
                            label="📊 Descargar Datos CSV",
                            data=csv_data,
                            file_name=f"datos_simulacion_{target_percentage*100:.1f}percent.csv",
                            mime="text/csv"
                        )

if __name__ == "__main__":
    main()